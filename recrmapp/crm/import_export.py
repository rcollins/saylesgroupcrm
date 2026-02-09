"""
CSV and Excel import/export for Lead, Client, Contact, Property.
Export: download as CSV or .xlsx.
Import: upload CSV or .xlsx, validate, create records (optional update by matching email/name).
"""
import csv
import io
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse

from .models import Lead, Client, Contact, Property

# Max upload size for import files (DoS prevention)
MAX_IMPORT_FILE_SIZE = 15 * 1024 * 1024  # 15 MB


# --- Export column definitions: (field_name, header_label) ---
EXPORT_COLUMNS = {
    'lead': [
        ('first_name', 'First Name'),
        ('last_name', 'Last Name'),
        ('email', 'Email'),
        ('phone', 'Phone'),
        ('referral', 'Referral'),
        ('status', 'Status'),
        ('address', 'Address'),
        ('city', 'City'),
        ('state', 'State'),
        ('zip_code', 'Zip Code'),
        ('notes', 'Notes'),
    ],
    'client': [
        ('first_name', 'First Name'),
        ('last_name', 'Last Name'),
        ('email', 'Email'),
        ('phone', 'Phone'),
        ('spouse_first_name', 'Spouse First Name'),
        ('spouse_last_name', 'Spouse Last Name'),
        ('spouse_email', 'Spouse Email'),
        ('spouse_phone', 'Spouse Phone'),
        ('client_type', 'Client Type'),
        ('status', 'Status'),
        ('address', 'Address'),
        ('city', 'City'),
        ('state', 'State'),
        ('zip_code', 'Zip Code'),
        ('budget_min', 'Budget Min'),
        ('budget_max', 'Budget Max'),
        ('notes', 'Notes'),
    ],
    'contact': [
        ('first_name', 'First Name'),
        ('last_name', 'Last Name'),
        ('email', 'Email'),
        ('phone', 'Phone'),
        ('contact_type', 'Contact Type'),
        ('company', 'Company'),
        ('address', 'Address'),
        ('city', 'City'),
        ('state', 'State'),
        ('zip_code', 'Zip Code'),
        ('notes', 'Notes'),
    ],
    'property': [
        ('title', 'Title'),
        ('property_type', 'Property Type'),
        ('status', 'Status'),
        ('address', 'Address'),
        ('city', 'City'),
        ('state', 'State'),
        ('zip_code', 'Zip Code'),
        ('price', 'Price'),
        ('bedrooms', 'Bedrooms'),
        ('bathrooms', 'Bathrooms'),
        ('square_feet', 'Square Feet'),
        ('lot_size', 'Lot Size'),
        ('year_built', 'Year Built'),
        ('mls_number', 'MLS Number'),
        ('mls_service', 'MLS Service'),
        ('mls_url', 'MLS URL'),
        ('description', 'Description'),
        ('features', 'Features'),
        ('featured', 'Featured'),
    ],
}


def _row_from_instance(instance, columns):
    """Build a list of cell values for one model instance."""
    row = []
    for field_name, _ in columns:
        val = getattr(instance, field_name, None)
        if val is None:
            row.append('')
        elif hasattr(val, 'isoformat'):  # date/datetime
            row.append(val.isoformat() if val else '')
        elif isinstance(val, (Decimal,)):
            row.append(str(val))
        elif isinstance(val, bool):
            row.append('Yes' if val else 'No')
        else:
            row.append(str(val))
    return row


def export_queryset_csv(queryset, columns, filename_base):
    """Stream queryset as CSV; return HttpResponse."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
    response.write('\ufeff')  # BOM for Excel UTF-8
    writer = csv.writer(response)
    headers = [label for _, label in columns]
    writer.writerow(headers)
    for obj in queryset:
        writer.writerow(_row_from_instance(obj, columns))
    return response


def export_queryset_xlsx(queryset, columns, filename_base):
    """Stream queryset as Excel .xlsx; return HttpResponse."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        from django.http import HttpResponse
        r = HttpResponse('Excel export requires openpyxl.', status=501)
        return r
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    headers = [label for _, label in columns]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for row_idx, obj in enumerate(queryset, 2):
        row_vals = _row_from_instance(obj, columns)
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if isinstance(val, str) and val.isdigit():
                try:
                    cell.value = int(val)
                except ValueError:
                    pass
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
    wb.save(response)
    return response


def _normalize_header(s):
    """Normalize header for matching: strip, lower, collapse spaces."""
    if not s:
        return ''
    return ' '.join(str(s).strip().lower().split())


def _parse_import_row(row, columns):
    """Convert a dict row (header -> value) into a dict of model field -> value. Match row keys to column labels."""
    out = {}
    for field_name, label in columns:
        norm_label = _normalize_header(label)
        norm_field = _normalize_header(field_name.replace('_', ' '))
        for h, val in row.items():
            norm_h = _normalize_header(h)
            if norm_h == norm_label or norm_h == norm_field:
                out[field_name] = val
                break
    return out


def _coerce_value(val, field_name, model_class):
    """Coerce string value to the right type for the model field."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, str):
        val = val.strip()
    try:
        field = model_class._meta.get_field(field_name)
    except Exception:
        return val
    if field.get_internal_type() in ('IntegerField', 'PositiveIntegerField', 'SmallIntegerField'):
        try:
            return int(float(str(val).replace(',', '')))
        except (ValueError, TypeError):
            return None
    if field.get_internal_type() in ('DecimalField',):
        try:
            return Decimal(str(val).replace(',', '').replace('$', '').strip())
        except (InvalidOperation, ValueError, TypeError):
            return None
    if field.get_internal_type() == 'BooleanField':
        if isinstance(val, bool):
            return val
        s = str(val).strip().lower()
        return s in ('1', 'true', 'yes', 'y', 'x')
    if field.get_internal_type() == 'URLField':
        s = str(val).strip()
        return s if s else ''
    return val


def _get_reader_for_file(uploaded_file, format_type):
    """Return (headers, row_iter). format_type is 'csv' or 'xlsx'. Raises ValueError if file too large."""
    if getattr(uploaded_file, 'size', 0) and uploaded_file.size > MAX_IMPORT_FILE_SIZE:
        raise ValueError(f'File too large. Maximum size is {MAX_IMPORT_FILE_SIZE // (1024 * 1024)} MB.')
    if format_type == 'csv':
        content = uploaded_file.read()
        if getattr(uploaded_file, 'size', 0) and len(content) > MAX_IMPORT_FILE_SIZE:
            raise ValueError(f'File too large. Maximum size is {MAX_IMPORT_FILE_SIZE // (1024 * 1024)} MB.')
        if hasattr(content, 'decode'):
            content = content.decode('utf-8-sig')  # strip BOM
        reader = csv.DictReader(io.StringIO(content))
        headers = reader.fieldnames or []
        return headers, reader
    if format_type == 'xlsx':
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError('Excel import requires openpyxl.')
        wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], iter([])
        headers = [str(c) if c is not None else '' for c in rows[0]]
        def row_iter():
            for row in rows[1:]:
                vals = [str(c).strip() if c is not None else '' for c in row]
                vals = vals + [''] * (len(headers) - len(vals))
                yield dict(zip(headers, vals))
        return headers, row_iter()
    raise ValueError(f'Unsupported format: {format_type}')


def import_records(uploaded_file, model_key, format_type):
    """
    Parse uploaded file and create records.
    model_key: 'lead' | 'client' | 'contact' | 'property'
    format_type: 'csv' | 'xlsx'
    Returns: dict with keys: created (int), errors (list of {row, message}).
    """
    model_map = {'lead': Lead, 'client': Client, 'contact': Contact, 'property': Property}
    model_class = model_map.get(model_key)
    if not model_class:
        return {'created': 0, 'errors': [{'row': 0, 'message': 'Invalid model.'}]}
    columns = EXPORT_COLUMNS.get(model_key, [])

    try:
        headers, row_iter = _get_reader_for_file(uploaded_file, format_type)
    except Exception as e:
        return {'created': 0, 'errors': [{'row': 0, 'message': str(e)}]}

    created = 0
    errors = []
    for row_num, row in enumerate(row_iter, 2):  # 2 = header is row 1
        if not isinstance(row, dict):
            row = dict(zip(headers, row)) if headers else {}
        data = _parse_import_row(row, columns)
        # Skip empty rows
        if all(not str(v).strip() for v in data.values() if v is not None):
            continue
        try:
            kwargs = {}
            for field_name, _ in columns:
                if field_name not in data:
                    continue
                raw = data.get(field_name)
                val = _coerce_value(raw, field_name, model_class)
                if model_key == 'property' and field_name == 'featured':
                    val = bool(val) if val is not None else False
                kwargs[field_name] = val
            # Required fields for each model
            if model_key == 'lead':
                if not kwargs.get('first_name') and not kwargs.get('last_name'):
                    errors.append({'row': row_num, 'message': 'First name or last name required.'})
                    continue
            if model_key == 'client':
                if not kwargs.get('first_name') and not kwargs.get('last_name'):
                    errors.append({'row': row_num, 'message': 'First name or last name required.'})
                    continue
            if model_key == 'contact':
                if not kwargs.get('first_name') and not kwargs.get('last_name'):
                    errors.append({'row': row_num, 'message': 'First name or last name required.'})
                    continue
            if model_key == 'property':
                if not kwargs.get('title'):
                    errors.append({'row': row_num, 'message': 'Title required.'})
                    continue
                if not kwargs.get('address'):
                    kwargs['address'] = kwargs.get('title', '')
            model_class.objects.create(**kwargs)
            created += 1
        except Exception as e:
            errors.append({'row': row_num, 'message': str(e)})
    return {'created': created, 'errors': errors}
